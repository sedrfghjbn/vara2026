from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils import timezone
from django.urls import reverse
from datetime import datetime, date
import csv
import json
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO

from .models import User, Employee, Vacancy, Training, Certificate, Department, Position
from hr.integrations.telegram.models import EmployeeTelegramLink, UserTelegramLink
from hr.integrations.telegram.linking import get_user_bind_code
from hr.integrations.telegram.dispatcher import EventDispatcher
from hr.integrations.telegram.formatter import build_employee_payload
from .forms import (
    RegisterForm, LoginForm, EmployeeForm, VacancyForm,
    TrainingForm, DepartmentForm, PositionForm, UserProfileForm
)


# ==================== Аутентификация ====================

def register_view(request):
    """Регистрация пользователя"""
    if request.user.is_authenticated:
        return redirect('hr:index')
    
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Автоматически входим после регистрации
            login(request, user)
            messages.success(request, f'Аккаунт создан! Добро пожаловать, {user.email}!')
            return redirect('hr:index')
    else:
        form = RegisterForm()
    
    return render(request, 'hr/register.html', {'form': form})


def login_view(request):
    """Вход в систему"""
    if request.user.is_authenticated:
        return redirect('hr:index')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email').lower()
            password = form.cleaned_data.get('password')
            # Находим пользователя по email и аутентифицируем его username-ом
            user_obj = User.objects.filter(email=email).first()
            if user_obj:
                user = authenticate(request, username=user_obj.username, password=password)
                if user is not None:
                    login(request, user)
                    messages.success(request, f'Добро пожаловать, {user.email}!')
                    return redirect('hr:index')
            messages.error(request, 'Неверный email или пароль.')
    else:
        initial_email = request.GET.get('email', '')
        form = LoginForm(initial={'email': initial_email})
    
    return render(request, 'hr/login.html', {'form': form})


@login_required
def logout_view(request):
    """Выход из системы"""
    logout(request)
    messages.success(request, 'Вы вышли из системы.')
    return redirect('hr:login')


@login_required
def profile_view(request):
    """Профиль пользователя: редактирование своих данных"""
    user = request.user
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Профиль обновлён.')
            return redirect('hr:profile')
    else:
        form = UserProfileForm(instance=user)

    employee = None
    try:
        employee = user.employee_profile
    except Employee.DoesNotExist:
        pass

    telegram_link = None
    if employee:
        telegram_link = EmployeeTelegramLink.objects.filter(
            employee=employee, is_active=True,
        ).first()
    if not telegram_link:
        telegram_link = UserTelegramLink.objects.filter(
            user=user, is_active=True,
        ).first()

    bind_code = get_user_bind_code(user)

    context = {
        'form': form,
        'employee': employee,
        'telegram_link': telegram_link,
        'telegram_bind_code': bind_code,
    }
    return render(request, 'hr/profile.html', context)


# ==================== Главная страница ====================

def index(request):
    """Главная страница (дашборд) или лендинг для гостей"""
    if not request.user.is_authenticated:
        return render(request, 'hr/landing.html')

    context = {}
    
    if request.user.is_hr_manager():
        # Статистика для HR-менеджеров
        context.update({
            'total_employees': Employee.objects.filter(status='active').count(),
            'total_vacancies': Vacancy.objects.filter(status='open').count(),
            'active_trainings': Training.objects.filter(status='in_progress').count(),
            'recent_employees': Employee.objects.order_by('-created_at')[:5],
            'recent_vacancies': Vacancy.objects.order_by('-created_at')[:5],
        })
    else:
        # Для обычных сотрудников - только свои обучения
        try:
            employee = request.user.employee_profile
            # Получаем все обучения, где сотрудник является участником
            my_trainings = Training.objects.filter(participants=employee).order_by('-start_date')
            context.update({
                'employee': employee,
                'my_trainings': my_trainings,
            })
        except Employee.DoesNotExist:
            pass
        except:
            pass
    
    return render(request, 'hr/index.html', context)


# ==================== Сотрудники ====================

@login_required
def employees_list(request):
    """Список сотрудников"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    employees = Employee.objects.all()
    
    # Поиск
    search_query = request.GET.get('search', '')
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Фильтры
    department_filter = request.GET.get('department', '')
    position_filter = request.GET.get('position', '')
    status_filter = request.GET.get('status', '')
    
    if department_filter:
        employees = employees.filter(department_id=department_filter)
    if position_filter:
        employees = employees.filter(position_id=position_filter)
    if status_filter:
        employees = employees.filter(status=status_filter)
    
    # Пагинация
    paginator = Paginator(employees, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'departments': Department.objects.all(),
        'positions': Position.objects.all(),
        'search_query': search_query,
        'department_filter': department_filter,
        'position_filter': position_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'hr/employees_list.html', context)


@login_required
def employee_detail(request, employee_id):
    """Детальная информация о сотруднике"""
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Проверка доступа
    if not request.user.is_hr_manager():
        try:
            if request.user.employee_profile.id != employee_id:
                messages.error(request, 'У вас нет доступа к этой странице.')
                return redirect('hr:index')
        except:
            messages.error(request, 'У вас нет доступа к этой странице.')
            return redirect('hr:index')
    
    trainings = Training.objects.filter(participants=employee)
    certificates = Certificate.objects.filter(employee=employee)
    
    context = {
        'employee': employee,
        'trainings': trainings,
        'certificates': certificates,
    }
    
    return render(request, 'hr/employee_detail.html', context)


@login_required
def employee_create(request):
    """Создание сотрудника"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    # Получаем доступных пользователей для модального окна
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            form.instance._updated_by_user = request.user
            employee = form.save()
            messages.success(request, f'Сотрудник {employee.full_name} успешно создан и связан с пользователем {employee.user.username}!')
            return redirect('hr:employee_detail', employee_id=employee.id)
    else:
        form = EmployeeForm()
        # Фильтруем только пользователей с ролью employee, у которых еще нет сотрудника
        form.fields['user'].queryset = User.objects.filter(
            role='employee'
        ).exclude(
            employee_profile__isnull=False
        ).order_by('username')
    
    # Получаем список доступных пользователей для модального окна
    available_users = User.objects.filter(role='employee').exclude(
        employee_profile__isnull=False
    ).order_by('username')
    
    context = {
        'form': form,
        'title': 'Добавить сотрудника',
        'available_users': available_users,
        'selected_user_id': None,
        'positions': Position.objects.all().order_by('name'),
        'departments': Department.objects.all().order_by('name'),
    }
    return render(request, 'hr/employee_form.html', context)


@login_required
def employee_edit(request, employee_id):
    """Редактирование сотрудника"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            form.instance._updated_by_user = request.user
            employee = form.save()
            messages.success(request, f'Информация о сотруднике {employee.full_name} обновлена!')
            return redirect('hr:employee_detail', employee_id=employee.id)
    else:
        form = EmployeeForm(instance=employee)
        form.fields['user'].queryset = User.objects.filter(
            role='employee'
        ).order_by('username')
    
    # Получаем список доступных пользователей для модального окна (включая текущего)
    available_users = User.objects.filter(role='employee').order_by('username')
    selected_user_id = employee.user.id if employee.user else None
    
    context = {
        'form': form,
        'employee': employee,
        'title': 'Редактировать сотрудника',
        'available_users': available_users,
        'selected_user_id': selected_user_id
    }
    return render(request, 'hr/employee_form.html', context)


@login_required
def employee_delete(request, employee_id):
    """Удаление сотрудника"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    employee = get_object_or_404(Employee, id=employee_id)
    
    if request.method == 'POST':
        employee_name = employee.full_name
        payload = build_employee_payload(
            employee,
            change_type='deleted',
            actor_email=request.user.email,
            actor_role='admin' if request.user.is_superuser else 'hr_manager' if request.user.role == 'hr_manager' else 'employee',
        )
        EventDispatcher.dispatch('employee_deleted', payload)
        employee.delete()
        messages.success(request, f'Сотрудник {employee_name} удален!')
        return redirect('hr:employees_list')
    
    return render(request, 'hr/employee_confirm_delete.html', {'employee': employee})


# ==================== Вакансии ====================

@login_required
def vacancies_list(request):
    """Список вакансий"""
    if request.user.is_hr_manager():
        # HR-менеджеры видят все вакансии
        vacancies = Vacancy.objects.all()
    else:
        # Обычные сотрудники видят только открытые вакансии
        vacancies = Vacancy.objects.filter(status='open')
    
    # Фильтр по статусу
    status_filter = request.GET.get('status', '')
    if status_filter:
        vacancies = vacancies.filter(status=status_filter)
    
    # Поиск
    search_query = request.GET.get('search', '')
    if search_query:
        vacancies = vacancies.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Пагинация
    paginator = Paginator(vacancies, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    
    return render(request, 'hr/vacancies_list.html', context)


@login_required
def vacancy_detail(request, vacancy_id):
    """Детальная информация о вакансии"""
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)
    
    # Обычные сотрудники могут видеть только открытые вакансии
    if not request.user.is_hr_manager() and vacancy.status != 'open':
        messages.error(request, 'У вас нет доступа к этой вакансии.')
        return redirect('hr:vacancies_list')
    
    return render(request, 'hr/vacancy_detail.html', {'vacancy': vacancy})


@login_required
def vacancy_create(request):
    """Создание вакансии"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    if request.method == 'POST':
        form = VacancyForm(request.POST)
        if form.is_valid():
            vacancy = form.save(commit=False)
            vacancy.hr_manager = request.user
            vacancy.save()
            messages.success(request, f'Вакансия "{vacancy.title}" успешно создана!')
            return redirect('hr:vacancy_detail', vacancy_id=vacancy.id)
    else:
        form = VacancyForm()
    
    context = {
        'form': form,
        'title': 'Создать вакансию',
        'positions': Position.objects.all().order_by('name'),
        'departments': Department.objects.all().order_by('name'),
    }
    return render(request, 'hr/vacancy_form.html', context)


@login_required
def vacancy_edit(request, vacancy_id):
    """Редактирование вакансии"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)
    
    if request.method == 'POST':
        form = VacancyForm(request.POST, instance=vacancy)
        if form.is_valid():
            vacancy = form.save()
            if vacancy.status == 'closed' and not vacancy.closed_at:
                vacancy.closed_at = timezone.now()
                vacancy.save()
            messages.success(request, f'Вакансия "{vacancy.title}" обновлена!')
            return redirect('hr:vacancy_detail', vacancy_id=vacancy.id)
    else:
        form = VacancyForm(instance=vacancy)
        if vacancy.position:
            form.fields['position_name'].initial = vacancy.position.name
        if vacancy.department:
            form.fields['department_name'].initial = vacancy.department.name
    
    context = {
        'form': form,
        'vacancy': vacancy,
        'title': 'Редактировать вакансию',
        'positions': Position.objects.all().order_by('name'),
        'departments': Department.objects.all().order_by('name'),
    }
    return render(request, 'hr/vacancy_form.html', context)


@login_required
def vacancy_delete(request, vacancy_id):
    """Удаление вакансии"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)
    
    if request.method == 'POST':
        vacancy_title = vacancy.title
        vacancy.delete()
        messages.success(request, f'Вакансия "{vacancy_title}" удалена!')
        return redirect('hr:vacancies_list')
    
    return render(request, 'hr/vacancy_confirm_delete.html', {'vacancy': vacancy})


# ==================== Обучение ====================

@login_required
def trainings_list(request):
    """Список обучения"""
    if request.user.is_hr_manager():
        # HR-менеджер видит все обучения
        trainings = Training.objects.all()
        
        # Фильтр по статусу
        status_filter = request.GET.get('status', '')
        if status_filter:
            trainings = trainings.filter(status=status_filter)
        
        # Поиск
        search_query = request.GET.get('search', '')
        if search_query:
            trainings = trainings.filter(
                Q(title__icontains=search_query) |
                Q(description__icontains=search_query)
            )
    else:
        # Обычный сотрудник видит только свои обучения
        try:
            employee = request.user.employee_profile
            trainings = Training.objects.filter(participants=employee)
            
            # Фильтр по статусу
            status_filter = request.GET.get('status', '')
            if status_filter:
                trainings = trainings.filter(status=status_filter)
            
            # Поиск
            search_query = request.GET.get('search', '')
            if search_query:
                trainings = trainings.filter(
                    Q(title__icontains=search_query) |
                    Q(description__icontains=search_query)
                )
        except:
            trainings = Training.objects.none()
            messages.info(request, 'Ваш профиль сотрудника не найден.')
    
    # Пагинация
    paginator = Paginator(trainings, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status_filter': request.GET.get('status', ''),
        'search_query': request.GET.get('search', ''),
    }
    
    return render(request, 'hr/trainings_list.html', context)


@login_required
def training_detail(request, training_id):
    """Детальная информация об обучении"""
    training = get_object_or_404(Training, id=training_id)
    
    # Проверка доступа
    if not request.user.is_hr_manager():
        # Обычный сотрудник может видеть только свои обучения
        try:
            employee = request.user.employee_profile
            if employee not in training.participants.all():
                messages.error(request, 'У вас нет доступа к этому обучению.')
                return redirect('hr:trainings_list')
        except:
            messages.error(request, 'Ваш профиль сотрудника не найден.')
            return redirect('hr:index')
    
    participants = training.participants.all()
    certificates = Certificate.objects.filter(training=training)
    
    context = {
        'training': training,
        'participants': participants,
        'certificates': certificates,
    }
    
    return render(request, 'hr/training_detail.html', context)


@login_required
def training_create(request):
    """Создание обучения - только для HR-менеджеров"""
    if not request.user.is_hr_manager():
        messages.error(request, 'Только HR-менеджеры могут создавать обучения.')
        return redirect('hr:index')
    
    # Получаем всех сотрудников с аккаунтами для модального окна
    available_employees = Employee.objects.filter(user__isnull=False).order_by('last_name', 'first_name')
    
    if request.method == 'POST':
        form = TrainingForm(request.POST)
        if form.is_valid():
            training = form.save(commit=False)
            if not training.responsible:
                training.responsible = request.user
            if training.status == 'in_progress':
                training._started_by_user = request.user
            training.save()
            form.save_m2m()  # Сохраняем many-to-many связи
            messages.success(request, f'Обучение "{training.title}" успешно создано!')
            return redirect('hr:training_detail', training_id=training.id)
    else:
        form = TrainingForm(initial={'responsible': request.user})
    
    context = {
        'form': form,
        'title': 'Создать обучение',
        'available_employees': available_employees,
        'selected_participants': []
    }
    return render(request, 'hr/training_form.html', context)


@login_required
def training_edit(request, training_id):
    """Редактирование обучения - только для HR-менеджеров"""
    if not request.user.is_hr_manager():
        messages.error(request, 'Только HR-менеджеры могут редактировать обучения.')
        return redirect('hr:index')
    
    training = get_object_or_404(Training, id=training_id)
    
    if request.method == 'POST':
        form = TrainingForm(request.POST, instance=training)
        if form.is_valid():
            old_status = training.status
            training = form.save(commit=False)
            if old_status != 'in_progress' and training.status == 'in_progress':
                training._started_by_user = request.user
            training.save()
            form.save_m2m()  # Сохраняем many-to-many связи (участников)
            messages.success(request, f'Обучение "{training.title}" обновлено!')
            return redirect('hr:training_detail', training_id=training.id)
    else:
        # Инициализируем форму с данными обучения
        form = TrainingForm(instance=training)
        # Убеждаемся, что даты в правильном формате для HTML5 date input
        if training.start_date:
            form.fields['start_date'].initial = training.start_date.strftime('%Y-%m-%d')
        if training.end_date:
            form.fields['end_date'].initial = training.end_date.strftime('%Y-%m-%d')
    
    # Получаем всех сотрудников с аккаунтами для модального окна
    available_employees = Employee.objects.filter(user__isnull=False).order_by('last_name', 'first_name')
    # Получаем уже выбранных участников
    selected_participants = list(training.participants.all().values_list('id', flat=True))
    
    context = {
        'form': form,
        'training': training,
        'title': 'Редактировать обучение',
        'available_employees': available_employees,
        'selected_participants': selected_participants
    }
    return render(request, 'hr/training_form.html', context)


@login_required
def training_delete(request, training_id):
    """Удаление обучения - только для HR-менеджеров"""
    if not request.user.is_hr_manager():
        messages.error(request, 'Только HR-менеджеры могут удалять обучения.')
        return redirect('hr:index')
    
    training = get_object_or_404(Training, id=training_id)
    
    if request.method == 'POST':
        training_title = training.title
        training.delete()
        messages.success(request, f'Обучение "{training_title}" удалено!')
        return redirect('hr:trainings_list')
    
    return render(request, 'hr/training_confirm_delete.html', {'training': training})


# ==================== Отчётность ====================

@login_required
def reports(request):
    """Страница с отчётами"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    return render(request, 'hr/reports.html')


@login_required
def report_vacancies(request):
    """Отчёт по вакансиям"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    format_type = request.GET.get('format', 'html')
    vacancies = Vacancy.objects.all().order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="vacancies_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Название', 'Статус', 'Отдел', 'Должность', 'Зарплата от', 'Зарплата до', 'HR-менеджер', 'Дата создания'])
        for v in vacancies:
            writer.writerow([
                v.title, v.get_status_display(), 
                v.department.name if v.department else '',
                v.position.name if v.position else '',
                v.salary_min or '', v.salary_max or '',
                v.hr_manager.username if v.hr_manager else '',
                v.created_at.strftime('%Y-%m-%d')
            ])
        return response
    
    elif format_type == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Вакансии"
        
        headers = ['Название', 'Статус', 'Отдел', 'Должность', 'Зарплата от', 'Зарплата до', 'HR-менеджер', 'Дата создания']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for v in vacancies:
            ws.append([
                v.title, v.get_status_display(),
                v.department.name if v.department else '',
                v.position.name if v.position else '',
                v.salary_min or '', v.salary_max or '',
                v.hr_manager.username if v.hr_manager else '',
                v.created_at.strftime('%Y-%m-%d')
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="vacancies_report.xlsx"'
        wb.save(response)
        return response
    
    elif format_type == 'json':
        data = [{
            'title': v.title,
            'status': v.get_status_display(),
            'department': v.department.name if v.department else None,
            'position': v.position.name if v.position else None,
            'salary_min': str(v.salary_min) if v.salary_min else None,
            'salary_max': str(v.salary_max) if v.salary_max else None,
            'hr_manager': v.hr_manager.username if v.hr_manager else None,
            'created_at': v.created_at.isoformat(),
        } for v in vacancies]
        
        response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="vacancies_report.json"'
        return response
    
    # HTML отчёт
    open_count = vacancies.filter(status='open').count()
    closed_count = vacancies.filter(status='closed').count()
    archived_count = vacancies.filter(status='archived').count()
    
    context = {
        'vacancies': vacancies,
        'open_count': open_count,
        'closed_count': closed_count,
        'archived_count': archived_count,
    }
    
    return render(request, 'hr/report_vacancies.html', context)


@login_required
def report_training(request):
    """Отчёт по обучению сотрудников"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    format_type = request.GET.get('format', 'html')
    trainings = Training.objects.all().order_by('-start_date')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="training_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Название', 'Статус', 'Ответственный', 'Дата начала', 'Дата окончания', 'Количество участников'])
        for t in trainings:
            writer.writerow([
                t.title, t.get_status_display(),
                t.responsible.username if t.responsible else '',
                t.start_date.strftime('%Y-%m-%d'),
                t.end_date.strftime('%Y-%m-%d'),
                t.participants.count()
            ])
        return response
    
    elif format_type == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Обучение"
        
        headers = ['Название', 'Статус', 'Ответственный', 'Дата начала', 'Дата окончания', 'Количество участников']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for t in trainings:
            ws.append([
                t.title, t.get_status_display(),
                t.responsible.username if t.responsible else '',
                t.start_date.strftime('%Y-%m-%d'),
                t.end_date.strftime('%Y-%m-%d'),
                t.participants.count()
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="training_report.xlsx"'
        wb.save(response)
        return response
    
    elif format_type == 'json':
        data = [{
            'title': t.title,
            'status': t.get_status_display(),
            'responsible': t.responsible.username if t.responsible else None,
            'start_date': t.start_date.isoformat(),
            'end_date': t.end_date.isoformat(),
            'participants_count': t.participants.count(),
        } for t in trainings]
        
        response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="training_report.json"'
        return response
    
    # HTML отчёт
    planned_count = trainings.filter(status='planned').count()
    in_progress_count = trainings.filter(status='in_progress').count()
    completed_count = trainings.filter(status='completed').count()
    
    context = {
        'trainings': trainings,
        'planned_count': planned_count,
        'in_progress_count': in_progress_count,
        'completed_count': completed_count,
    }
    
    return render(request, 'hr/report_training.html', context)


@login_required
def report_staffing(request):
    """Штатное расписание"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    format_type = request.GET.get('format', 'html')
    employees = Employee.objects.filter(status='active').order_by('department', 'position', 'last_name')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="staffing_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['ФИО', 'Должность', 'Отдел', 'Email', 'Телефон', 'Дата приема'])
        for e in employees:
            writer.writerow([
                e.full_name, e.position.name if e.position else '',
                e.department.name if e.department else '',
                e.email, e.phone, e.hire_date.strftime('%Y-%m-%d')
            ])
        return response
    
    elif format_type == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Штатное расписание"
        
        headers = ['ФИО', 'Должность', 'Отдел', 'Email', 'Телефон', 'Дата приема']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for e in employees:
            ws.append([
                e.full_name, e.position.name if e.position else '',
                e.department.name if e.department else '',
                e.email, e.phone, e.hire_date.strftime('%Y-%m-%d')
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="staffing_report.xlsx"'
        wb.save(response)
        return response
    
    elif format_type == 'json':
        data = [{
            'full_name': e.full_name,
            'position': e.position.name if e.position else None,
            'department': e.department.name if e.department else None,
            'email': e.email,
            'phone': e.phone,
            'hire_date': e.hire_date.isoformat(),
        } for e in employees]
        
        response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="staffing_report.json"'
        return response
    
    # HTML отчёт
    # Группировка по отделам
    departments_data = {}
    for emp in employees:
        dept_name = emp.department.name if emp.department else 'Без отдела'
        if dept_name not in departments_data:
            departments_data[dept_name] = []
        departments_data[dept_name].append(emp)
    
    context = {
        'employees': employees,
        'departments_data': departments_data,
        'total_count': employees.count(),
    }
    
    return render(request, 'hr/report_staffing.html', context)


# ==================== Импорт данных ====================

@login_required
def import_employees(request):
    """Импорт сотрудников из CSV/XLSX"""
    if not request.user.is_hr_manager():
        messages.error(request, 'У вас нет доступа к этой странице.')
        return redirect('hr:index')
    
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        imported_count = 0
        errors = []
        
        try:
            if file_extension == 'csv':
                # Импорт из CSV
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                
                for row in reader:
                    try:
                        # Получаем или создаем отдел
                        department = None
                        if row.get('department'):
                            department, _ = Department.objects.get_or_create(name=row['department'])
                        
                        # Получаем или создаем должность
                        position = None
                        if row.get('position'):
                            position, _ = Position.objects.get_or_create(name=row['position'])
                        
                        from hr.forms import validate_employee_birth_date
                        dob = (
                            datetime.strptime(row['date_of_birth'], '%Y-%m-%d').date()
                            if row.get('date_of_birth')
                            else date.today()
                        )
                        validate_employee_birth_date(dob)

                        # Создаем сотрудника
                        employee = Employee.objects.create(
                            first_name=row.get('first_name', ''),
                            last_name=row.get('last_name', ''),
                            middle_name=row.get('middle_name', ''),
                            date_of_birth=dob,
                            email=row.get('email', ''),
                            phone=row.get('phone', ''),
                            department=department,
                            position=position,
                            hire_date=datetime.strptime(row['hire_date'], '%Y-%m-%d').date() if row.get('hire_date') else date.today(),
                            status=row.get('status', 'active'),
                        )
                        imported_count += 1
                    except Exception as e:
                        errors.append(f"Ошибка в строке {row}: {str(e)}")
            
            elif file_extension in ['xlsx', 'xls']:
                # Импорт из Excel
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                
                # Читаем заголовки
                headers = [cell.value for cell in ws[1]]
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_dict = dict(zip(headers, row))
                        
                        # Получаем или создаем отдел
                        department = None
                        if row_dict.get('department'):
                            department, _ = Department.objects.get_or_create(name=str(row_dict['department']))
                        
                        # Получаем или создаем должность
                        position = None
                        if row_dict.get('position'):
                            position, _ = Position.objects.get_or_create(name=str(row_dict['position']))
                        
                        from hr.forms import validate_employee_birth_date
                        raw_dob = row_dict.get('date_of_birth')
                        if isinstance(raw_dob, date):
                            dob = raw_dob
                        elif raw_dob:
                            dob = datetime.strptime(str(raw_dob), '%Y-%m-%d').date()
                        else:
                            dob = date.today()
                        validate_employee_birth_date(dob)

                        # Создаем сотрудника
                        employee = Employee.objects.create(
                            first_name=str(row_dict.get('first_name', '')),
                            last_name=str(row_dict.get('last_name', '')),
                            middle_name=str(row_dict.get('middle_name', '')),
                            date_of_birth=dob,
                            email=str(row_dict.get('email', '')),
                            phone=str(row_dict.get('phone', '')),
                            department=department,
                            position=position,
                            hire_date=row_dict['hire_date'] if isinstance(row_dict.get('hire_date'), date) else datetime.strptime(str(row_dict['hire_date']), '%Y-%m-%d').date() if row_dict.get('hire_date') else date.today(),
                            status=str(row_dict.get('status', 'active')),
                        )
                        imported_count += 1
                    except Exception as e:
                        errors.append(f"Ошибка в строке {row_num}: {str(e)}")
            
            else:
                messages.error(request, 'Неподдерживаемый формат файла. Используйте CSV или XLSX.')
                return redirect('hr:employees_list')
            
            if imported_count > 0:
                messages.success(request, f'Успешно импортировано {imported_count} сотрудников.')
            if errors:
                messages.warning(request, f'Обнаружено {len(errors)} ошибок при импорте.')
        
        except Exception as e:
            messages.error(request, f'Ошибка при импорте: {str(e)}')
        
        return redirect('hr:employees_list')
    
    return render(request, 'hr/import_employees.html')

